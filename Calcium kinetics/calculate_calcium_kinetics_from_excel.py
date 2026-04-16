import csv
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


WORKBOOK = Path("Analysis_20250326_210022.xlsx")
SHEET = "Sheet1"
GROUPS = ["siScr", "siYAP", "siTAZ", "siYAPsiTAZ"]
OUTPUT_CSV = Path("ca_kinetics_per_replicate.csv")


def load_traces():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    trace_names = list(header[1:13])

    time = []
    traces = {name: [] for name in trace_names}
    for row in rows[2:]:
        if row[0] is None:
            continue
        values = row[1:13]
        if any(v is None for v in values):
            continue
        time.append(float(row[0]))
        for name, value in zip(trace_names, values):
            traces[name].append(float(value))
    return np.array(time), {name: np.array(vals) for name, vals in traces.items()}


def extrema(trace, mode="peak", min_gap=5):
    indices = []
    last_index = -999
    for i in range(1, len(trace) - 1):
        is_extreme = (
            trace[i] > trace[i - 1] and trace[i] >= trace[i + 1]
            if mode == "peak"
            else trace[i] < trace[i - 1] and trace[i] <= trace[i + 1]
        )
        if is_extreme and i - last_index >= min_gap:
            indices.append(i)
            last_index = i
    return indices


def trace_metrics(time, trace):
    peaks = extrema(trace, "peak", 5)
    troughs = extrema(trace, "trough", 5)
    rise_slopes = []
    decay_slopes = []
    time_to_peak = []
    decay_tau = []

    for peak_idx in peaks:
        previous_troughs = [idx for idx in troughs if idx < peak_idx]
        next_troughs = [idx for idx in troughs if idx > peak_idx]
        if not previous_troughs or not next_troughs:
            continue

        trough_before = previous_troughs[-1]
        trough_after = next_troughs[0]
        if time[peak_idx] == time[trough_before] or time[trough_after] == time[peak_idx]:
            continue

        rise_slopes.append((trace[peak_idx] - trace[trough_before]) / (time[peak_idx] - time[trough_before]))
        decay_slopes.append((trace[trough_after] - trace[peak_idx]) / (time[trough_after] - time[peak_idx]))
        time_to_peak.append(time[peak_idx] - time[trough_before])

        target = trace[trough_after] + (trace[peak_idx] - trace[trough_after]) / np.e
        decay_segment = trace[peak_idx : trough_after + 1]
        time_segment = time[peak_idx : trough_after + 1]
        crossing = np.where(decay_segment <= target)[0]
        if len(crossing):
            decay_tau.append(time_segment[crossing[0]] - time[peak_idx])

    return {
        "n_events": len(rise_slopes),
        "rise_slope_mean": float(np.mean(rise_slopes)),
        "decay_slope_mean": float(np.mean(decay_slopes)),
        "time_to_peak_mean_s": float(np.mean(time_to_peak)),
        "decay_tau_mean_s": float(np.mean(decay_tau)) if decay_tau else float("nan"),
    }


def main():
    time, traces = load_traces()
    by_replicate = {name: trace_metrics(time, trace) for name, trace in traces.items()}

    with OUTPUT_CSV.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "replicate",
                "group",
                "n_events",
                "rise_slope_mean",
                "decay_slope_mean",
                "time_to_peak_mean_s",
                "decay_tau_mean_s",
            ]
        )
        for replicate, metrics in by_replicate.items():
            writer.writerow(
                [
                    replicate,
                    replicate.split("-")[0],
                    metrics["n_events"],
                    metrics["rise_slope_mean"],
                    metrics["decay_slope_mean"],
                    metrics["time_to_peak_mean_s"],
                    metrics["decay_tau_mean_s"],
                ]
            )

    print("Per-replicate kinetics")
    for replicate, metrics in by_replicate.items():
        print(replicate, metrics)

    print("\nGroup summary")
    for group in GROUPS:
        replicates = [f"{group}-1", f"{group}-2", f"{group}-3"]
        rise = np.array([by_replicate[name]["rise_slope_mean"] for name in replicates])
        decay = np.array([by_replicate[name]["decay_slope_mean"] for name in replicates])
        time_to_peak = np.array([by_replicate[name]["time_to_peak_mean_s"] for name in replicates])
        tau = np.array([by_replicate[name]["decay_tau_mean_s"] for name in replicates])
        print(group)
        print(f"  rise slope: {rise.mean():.4f} +/- {rise.std(ddof=1):.4f}")
        print(f"  decay slope: {decay.mean():.4f} +/- {decay.std(ddof=1):.4f}")
        print(f"  time to peak (s): {time_to_peak.mean():.4f} +/- {time_to_peak.std(ddof=1):.4f}")
        print(f"  decay tau (s): {tau.mean():.4f} +/- {tau.std(ddof=1):.4f}")

    print(f"\nWrote {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
